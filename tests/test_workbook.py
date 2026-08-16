"""The headless workbook reader (openpyxl grid path) on a synthetic .xlsx.

Guards the grid logic — sheet/label lookup, the spacer-column value read, the
abbreviated-month ETD, and the VGM container table — without Excel/COM. The .xls
(xlrd) path is validated against live workbooks; here we cover the field logic.
"""

import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import sandbox  # noqa: F401 - keeps the real bootstrap pointer untouched

import openpyxl

from bot_kalung.services.workbook import read_shipment_fields

failures = []


def check(label, cond):
    print(f"{'PASS' if cond else 'FAIL'}  {label}")
    if not cond:
        failures.append(label)


with tempfile.TemporaryDirectory() as tmp:
    folder = Path(tmp) / "16.NIT"
    folder.mkdir()

    wb = openpyxl.Workbook()
    si = wb.active
    si.title = "SI  benar"
    # Destination with a spacer column: label in col 2, value in col 4.
    si.cell(row=5, column=2, value="Port of Discharge")
    si.cell(row=5, column=4, value="KARACHI, PAKISTAN")
    si.cell(row=6, column=2, value="ETD")
    si.cell(row=6, column=3, value="15 AUG 2026")          # abbreviated month
    si.cell(row=7, column=2, value="Party")
    si.cell(row=7, column=3, value="5 X 40'HC")

    vgm = wb.create_sheet("VGM ")
    vgm.cell(row=11, column=2, value="VESSEL NAME")
    vgm.cell(row=11, column=5, value="INTEGRA-184E")
    vgm.cell(row=12, column=2, value="BOOKING NO")
    vgm.cell(row=12, column=5, value="2318229000")
    vgm.cell(row=14, column=4, value="CONTAINER NO")       # table header
    vgm.cell(row=15, column=2, value=1)
    vgm.cell(row=15, column=3, value="40'HC")
    vgm.cell(row=15, column=4, value="cmau8513405")
    vgm.cell(row=16, column=2, value=2)
    vgm.cell(row=16, column=3, value="40'HC")
    vgm.cell(row=16, column=4, value="TRHU5986693")
    wb.save(folder / "NIT16-KATT-VGM,SI,INV,PL.xlsx")

    f = read_shipment_fields(folder)
    check("destination read across a spacer column",
          f.destination_port == "Karachi" and f.destination_country == "Pakistan")
    check("abbreviated-month ETD parsed", f.etd == date(2026, 8, 15))
    check("party size read",
          f.container_quantity == 5 and f.container_size_short == "40'HC")
    check("vessel + voyage read",
          f.vessel_name == "INTEGRA" and f.voyage == "184E")
    check("booking read", f.booking_number == "2318229000")
    check("containers read and upper-cased",
          f.containers == ["CMAU8513405", "TRHU5986693"])
    check("a good workbook reads without warnings", f.warnings == [])

    empty = Path(tmp) / "empty"
    empty.mkdir()
    f2 = read_shipment_fields(empty)
    check("a folder with no workbook warns instead of crashing",
          f2.vessel_name is None and bool(f2.warnings))

print()
if failures:
    print(f"{len(failures)} FAILED: {failures}")
    sys.exit(1)
print("Workbook OK - all checks passed.")
