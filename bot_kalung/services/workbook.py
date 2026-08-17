"""Headless workbook reading — the PC-free ingest path.

`excel.read_shipment_fields` reads a shipment's SI/VGM workbook through Excel COM
(xlwings), which needs Windows + installed Excel. This is the same read with no
Excel at all, so the folder scan can run on a plain Linux worker:

* `.xlsx` is read with openpyxl, using each formula cell's Excel-cached value
  (`data_only=True`) — verified to match the COM reader across every live shipment.
* `.xls` (older binary) is read with xlrd.

Both formats are normalised to a 1-indexed `_Grid` (cell values + merged spans),
so the field logic is identical and the same as `excel`'s. Only the loading
differs; the parsing (destination, ETD, party, vessel/voyage, booking) is the
exact same pure logic, reused from `excel`.
"""

from __future__ import annotations

import io
from pathlib import Path

from .excel import (
    ExcelError,
    ShipmentFields,
    _clean_booking,
    _normalize,
    _parse_long_date,
    _parse_party,
    _split_destination,
    _split_vessel_voyage,
    find_main_workbook,
)


# -- a format-agnostic worksheet grid ----------------------------------------

class _Grid:
    """One worksheet as {(row, col): value} (1-indexed) plus its merged spans,
    each merge a (min_row, max_row, min_col, max_col) inclusive tuple."""

    def __init__(self, name, cells, merged):
        self.name = name
        self._cells = cells
        self.merged = merged

    def cell(self, row, col):
        return self._cells.get((row, col))

    def ordered_cells(self):
        # Row-major, so a label search returns the same first match as Excel.
        return sorted(self._cells.items())


def _grids_from_xlsx(data: bytes) -> dict[str, _Grid]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    try:
        grids: dict[str, _Grid] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            cells = {(c.row, c.column): c.value
                     for row in ws.iter_rows() for c in row
                     if c.value is not None}
            merged = [(m.min_row, m.max_row, m.min_col, m.max_col)
                      for m in ws.merged_cells.ranges]
            grids[name] = _Grid(name, cells, merged)
        return grids
    finally:
        wb.close()


def _grids_from_xls(data: bytes) -> dict[str, _Grid]:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    grids: dict[str, _Grid] = {}
    for name in book.sheet_names():
        sh = book.sheet_by_name(name)
        cells = {}
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ctype = sh.cell_type(r, c)
                if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                value = sh.cell_value(r, c)
                if ctype == xlrd.XL_CELL_DATE:      # serial float -> datetime
                    value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                cells[(r + 1, c + 1)] = value
        # xlrd merges are (rlo, rhi, clo, chi), 0-indexed and half-open.
        merged = [(rlo + 1, rhi, clo + 1, chi)
                  for (rlo, rhi, clo, chi) in sh.merged_cells]
        grids[name] = _Grid(name, cells, merged)
    return grids


def _load_grids(source) -> dict[str, _Grid]:
    """Load a workbook's grids from a filesystem path or a Path-like Drive node.

    Reads the bytes once (Path.read_bytes / DriveNode.read_bytes) and parses from
    memory, so the same reader serves the local scan and the Drive-API scan.
    """
    name = source.name if hasattr(source, "name") else str(source)
    data = (source.read_bytes() if hasattr(source, "read_bytes")
            else Path(source).read_bytes())
    return (_grids_from_xls(data) if name.lower().endswith(".xls")
            else _grids_from_xlsx(data))


# -- cell lookup (mirrors excel.find_sheet / find_label / ...) ----------------

def _find_sheet(grids, prefix: str):
    """Match by normalized prefix so 'SI ', 'SI  ' and 'SI  benar' all resolve."""
    target = _normalize(prefix)
    for name, grid in grids.items():
        if _normalize(name).startswith(target):
            return grid
    return None


def _find_label(grid, text: str, *, column: int | None = None,
                exact: bool = False) -> tuple[int, int] | None:
    """Locate a label cell, case-insensitively. Returns (row, col) 1-based."""
    needle = text.strip().upper()
    for (row, col), value in grid.ordered_cells():
        if not isinstance(value, str):
            continue
        cell_text = value.strip().upper()
        if not cell_text:
            continue
        if column is not None and col != column:
            continue
        if (cell_text == needle) if exact else (needle in cell_text):
            return row, col
    return None


def _value_after(grid, row: int, col: int) -> int:
    """First column right of a label, skipping the label's merged span (some
    templates merge the label across two columns, e.g. NIT's ETD in G:H)."""
    for (min_row, max_row, min_col, max_col) in grid.merged:
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return max_col + 1
    return col + 1


def _cell(grid, row: int, col: int):
    return grid.cell(row, col)


def _value_right(grid, row: int, col: int, span: int = 3):
    """The value to the right of a label: the first non-empty cell after the
    label's merged span, within `span` columns. Some templates (e.g. MA's .xls)
    leave a spacer column between the label and its value, so col+1 is empty."""
    start = _value_after(grid, row, col)
    for c in range(start, start + span):
        value = grid.cell(row, c)
        if value is not None and (not isinstance(value, str) or value.strip()):
            return value
    return None


def _container_rows(grid) -> tuple[int, list[int]]:
    """(header_row, data_rows) for the VGM sheet's container table."""
    header = _find_label(grid, "CONTAINER NO")
    if header is None:
        raise ExcelError("Tidak menemukan tabel kontainer di sheet VGM.")
    header_row = header[0]
    rows: list[int] = []
    row = header_row + 1
    while True:
        index_value = _cell(grid, row, 2)
        if isinstance(index_value, (int, float)) and not isinstance(index_value, bool):
            rows.append(row)
            row += 1
            continue
        break
    return header_row, rows


# -- field reads (same logic as excel._read_si / _read_vgm) ------------------

def _read_si(grid, fields: ShipmentFields) -> None:
    for label in ("PORT OF DISCHARGE", "DESTINATION"):
        pod = _find_label(grid, label)
        if not pod:
            continue
        port, country = _split_destination(_value_right(grid, *pod))
        if port:
            fields.destination_port, fields.destination_country = port, country
            break

    etd = _find_label(grid, "ETD", exact=True)
    if etd:
        fields.etd = _parse_long_date(_value_right(grid, *etd))

    party = _find_label(grid, "Party", exact=True)
    if party:
        fields.container_quantity, fields.container_size_short = _parse_party(
            _value_right(grid, *party))


def _read_vgm(grid, fields: ShipmentFields) -> None:
    vessel = _find_label(grid, "VESSEL NAME", column=2)
    if vessel:
        fields.vessel_name, fields.voyage = _split_vessel_voyage(
            _cell(grid, vessel[0], 5))

    booking = _find_label(grid, "BOOKING NO", column=2)
    if booking:
        fields.booking_number = _clean_booking(_cell(grid, booking[0], 5))

    try:
        _, rows = _container_rows(grid)
    except ExcelError:
        fields.warnings.append("Tabel kontainer VGM tidak ditemukan.")
        return
    numbers = []
    for row in rows:
        value = _cell(grid, row, 4)      # column D — container number
        if isinstance(value, str) and value.strip():
            numbers.append(value.strip().upper())
    fields.containers = numbers
    if fields.container_quantity is None and rows:
        fields.container_quantity = len(rows)
    if fields.container_size_short is None and rows:
        size = _cell(grid, rows[0], 3)
        if isinstance(size, str) and size.strip():
            fields.container_size_short = size.strip()


def read_shipment_fields(folder) -> ShipmentFields:
    """Headless twin of `excel.read_shipment_fields` (.xlsx + .xls). Never raises
    — problems land in `warnings`."""
    fields = ShipmentFields()
    workbook = find_main_workbook(folder)
    if workbook is None or Path(workbook).name.startswith("~$"):
        fields.warnings.append("Tidak ada workbook VGM/SI/Inv/PL di folder ini.")
        return fields
    fields.workbook = workbook.name
    try:
        grids = _load_grids(workbook)
        si = _find_sheet(grids, "SI")
        if si is not None:
            _read_si(si, fields)
        else:
            fields.warnings.append("Sheet SI tidak ditemukan.")
        vgm = _find_sheet(grids, "VGM")
        if vgm is not None:
            _read_vgm(vgm, fields)
        else:
            fields.warnings.append("Sheet VGM tidak ditemukan.")
    except Exception as exc:      # noqa: BLE001 - a bad workbook must not crash a scan
        fields.warnings.append(f"Gagal membaca Excel: {exc}")
    return fields
